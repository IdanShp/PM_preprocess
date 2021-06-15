import sys
import csv
import os
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo




def create_fields():
    mat = []
    cols = ['A', 'B', 'C', 'D']
    rows = ['1', '2', '3']
    for c in cols:
        for r in rows:
            mat.append(c + r)
    return mat


def create_key_list(csv_rows, field_name):
    keys = set()
    for r in csv_rows:
        keys.add(r[field_name])
    sorted = list(keys)
    sorted.sort()
    return sorted


def make_table(workbook, sheet_name, key_list, matrix, cell_value_func,
               col_row_divisor="From  \\  To", info_text="this table mean somthing", info_text_cell="A1",
               table_start_row=2, table_start_col=2,
               table_style=TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True,
                                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)):

    workbook.create_sheet(sheet_name)
    sheet = workbook[sheet_name]
    sheet[info_text_cell] = info_text
    table_row_number = table_start_row
    table_col_number = table_start_col
    table_range = sheet.cell(row=table_row_number, column=table_col_number).coordinate + ':' \
                  + sheet.cell(row=table_row_number + len(key_list), column=table_col_number + len(key_list)).coordinate

    # print titles:
    row_number = table_row_number
    column_number = table_col_number
    sheet.cell(row=row_number, column=column_number).value = col_row_divisor
    row_number = row_number + 1

    # row titles
    empty_title_count = 1
    for f in key_list:
        t = str(f)
        if t == "":
            t = "empty " + str(empty_title_count)
            empty_title_count = empty_title_count + 1
        sheet.cell(row=row_number, column=column_number).value = t
        row_number = row_number + 1

    row_number = table_row_number
    column_number = table_col_number
    column_number = column_number + 1

    # column titles
    empty_title_count = 1
    for f in key_list:
        t = str(f)
        if t == "":
            t = "empty " + str(empty_title_count)
            empty_title_count = empty_title_count + 1
        sheet.cell(row=row_number, column=column_number).value = t
        column_number = column_number + 1

    # fill table:
    row_number = table_row_number + 1
    column_number = table_col_number + 1
    for i in range(len(matrix)):
        for j in range(len(matrix[i])):
            cell_value = cell_value_func(matrix, i, j)
            sheet.cell(row=row_number, column=column_number).value = cell_value
            column_number = column_number + 1
        row_number = row_number + 1
        column_number = table_col_number + 1

    # design table
    statis_tab = Table(displayName=sheet_name + "_table", ref=table_range)
    statis_tab.tableStyleInfo = table_style
    sheet.add_table(statis_tab)


def transitions_func(matrix, i, j):
    return matrix[i][j]

def statistic_func(matrix, i, j):
    row_sum = sum(matrix[i])
    cell_value = matrix[i][j] / row_sum if (row_sum != 0) else 0
    return cell_value

def dependency_func(matrix, i, j):
    numerator = (matrix[i][j] - matrix[j][i]) if (i != j) else (matrix[i][j])
    denominator = (matrix[i][j] + matrix[j][i] + 1) if (i != j) else (matrix[i][j] + 1)
    cell_value = numerator / denominator
    return cell_value


workbook = Workbook()
sheet = workbook.active

# Add a default style with striped rows and banded columns
tables_style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=True,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)

if len(sys.argv) < 2:
    print("you need to enter file path!")
    sys.exit()
    
if  not os.path.exists(sys.argv[1]):
    print("file dont exist")
    sys.exit()


# ------------------- read csv file into rows --------------
with open(str(sys.argv[1]), newline='') as csvfile:
    spamreader = csv.DictReader(csvfile, delimiter=',', quotechar='\"')
    col_number = 5
    rows = []
    for row in spamreader:
        rows.append(row)

# ------------------- fill excel - cases --------------

# define variables:
key_csv_name = 'player_name'
case_id_name = 'caseId'
# key_list = create_fields()
key_list = create_key_list(rows, key_csv_name)

# build 0 matrix
matrix = []
for out in range(0, len(key_list)):
    matrix_rox = []
    for inner in range(0, len(key_list)):
        matrix_rox.append(0)
    matrix.append(matrix_rox)

# print row titles:
sheet["A1"] = "Case Id \\ Event No."
i = 1
for c in range(2, 20):
    sheet.cell(row=1, column=c).value = str(c - 1)

# fill case data
i = 0
prev_case = -1
column_number = 1
row_number = 1
while i < (len(rows) - 1):
    # fill the case rows
    if rows[i][case_id_name] != prev_case:
        prev_case = rows[i][case_id_name]
        column_number = 1
        row_number = row_number + 1
        sheet.cell(row=row_number, column=column_number).value = rows[i][case_id_name]
        column_number = column_number + 1

    sheet.cell(row=row_number, column=column_number).value = rows[i][key_csv_name]
    column_number = column_number + 1
    i+=1

    # fill transitions:
    if rows[i][case_id_name] == prev_case:
        matrix[key_list.index(rows[i - 1][key_csv_name])][key_list.index(rows[i][key_csv_name])] = \
            matrix[key_list.index(rows[i - 1][key_csv_name])][key_list.index(rows[i][key_csv_name])] + 1

# make it a table with design
case_tab = Table(displayName="case_table", ref="A1:R60")
case_tab.tableStyleInfo = tables_style
sheet.add_table(case_tab)

# ------------------- fill excel - transitions --------------

# ---------make table test --------------
make_table(workbook, "transitions", key_list=key_list,matrix=matrix, cell_value_func=transitions_func, info_text="Number of times the ball moved from field  \'From\', the field \'To\' is as mentioned in the table")

make_table(workbook, "statistics", key_list,matrix, statistic_func, info_text="If the current field is \'From\', the chance to move to filed \'To\' is as mentioned in the table")

make_table(workbook, "dependency", key_list,matrix, dependency_func, info_text="dependency between  \'From\', to  \'To\' is as mentioned in the table")



if os.path.exists("excel_data.xlsx"):
    print("deleting old file " + "excel_data.xlsx")
    os.remove("excel_data.xlsx")
workbook.save(filename="excel_data_player_good.xlsx")
